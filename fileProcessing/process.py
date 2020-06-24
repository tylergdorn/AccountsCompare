"""Process is where we load all files into a standard format"""

import csv
from openpyxl import load_workbook
from typing import List, Union, Optional, Dict

from fileProcessing import classes
from fileProcessing import errors

class qbFormat:
    """Holder for the format of our quickbooks file"""
    def __init__(self, num: int, date: int, name: int, amount: int, account: int):
        self.num = num
        self.date = date
        self.name = name
        self.amount = amount
        self.account = account

def findQBHeader(indexRow: List) -> qbFormat: #type: ignore
    """ When given the first row in a quickbooks export, it gives us a qbFormat of where to find all the data in each row"""
    indexes: List[Optional[int]] = [None] * 5
    for index, item in enumerate(indexRow):
        if item.value == "Num":
            indexes[0] = index
        elif item.value == "Date":
            indexes[1] = index
        elif item.value == "Name":
            indexes[2] = index
        elif item.value == "Amount":
            indexes[3] = index
        elif item.value == "Account":
            indexes[4] = index
    for item in indexes:
        # if any of them are still None we raise an error
        if not item:
            raise errors.FileLoadError("Failed to load quickbooks header")
    return qbFormat(*indexes) #type: ignore

@errors.FileLoadDecorator
def loadAlProCSV(filePath: str) -> List[classes.Record]:
    """
    This loads an AL-Pro CSV file with '\t' as the delimiter. We may need to change this to accommodate better cases.
    This returns an array of Records
    """
    res = []
    with open(filePath, 'r', encoding="utf8", errors="replace") as csvfile:
        csvReader = csv.DictReader(csvfile, delimiter='\t')
        i = 1
        for row in csvReader:
            i += 1 # counting the line numbers the lazy way
            # sometimes this field is blank. why? no clue. 
            # This is the amount due
            try:
                num = float(row["TotalDue"])
            except:
                num = 0
            # if it's less than zero we dont' care
            if num > 0:
                # pass it all into Record if it's not negative
                record = classes.Record(row["InvoiceNo"], row["InvoiceDate"], row["AgencyName"], row["TotalDue"] if row["TotalDue"] else "0", i, True)
                res.append(record)
    csvfile.close()
    return res

@errors.FileLoadDecorator
def loadQBFile(filePath: str) -> List[classes.Record]:
    """
    This method loads a given file and returns an array of Records. A lot of this is hardcoded in so if the format of the QB file changes we will need to fix that. 
    We might want to think of a better way to do this.
    """
    wb = load_workbook(filename=filePath, read_only=True)
    ws = wb['Sheet1']
    consolidate: Dict[str, classes.Record] = {}
    ws.calculate_dimension(force=True)
    # we find the rows we want
    header = findQBHeader(next(ws.rows))
    for index, item in enumerate(ws.rows):
        # iterate through all rows and load them into our array
        # minus two because of the annoying sum at the bottom
        if 1 <= index < (ws.max_row - 3):
            record = _loadRow(item, index, header)
            if record:
                # we just store the last 5 digits. this is typically good enough
                record.invoiceNo = record.invoiceNo[:5]
                if record.invoiceNo in consolidate:
                    # if there is a collision we add the line number for usability and increment the totaldue amount
                    consolidate[record.invoiceNo].totalDue += record.totalDue
                    consolidate[record.invoiceNo].line.extend(record.line)
                else:
                    consolidate[record.invoiceNo] = record
        # make a list out of the values so we can iterate
    wb.close()
    return list(consolidate.values())
    
def _loadRow(row: List, rowNo: int, head: qbFormat) -> Union[classes.Record, None]: #type: ignore
    """_loadRow takes a row from a openpyxl workbook and returns a Record corresponding to the row. the Row no is passed along just to build the Record"""
    """ Returns None if we don't care about the row in question"""
    # if this is accounts receivable we return none. else we use the header data to tell us where to get the info
    try:
        return classes.Record(row[head.num].value, row[head.date].value, row[head.name].value, row[head.amount].value, rowNo, False) if "Accounts Receivable" in row[head.account].value else None
    except TypeError as e:
        # this is pretty lazy, but we get type errors when the row is empty. this should make this work more often
        return None
